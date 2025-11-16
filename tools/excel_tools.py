import csv
import os
from typing import Dict, List, Optional

import openpyxl


def register_excel_tools() -> Dict[str, dict]:
    return {
        "read_excel": {
            "description": "读取 Excel 文件（支持多 sheet）并返回结构化结果",
            "parameters": {
                "path": {"type": "string", "description": "Excel 文件路径"},
                "sheets": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "可选：只读取的工作表名称列表，为空则读取全部",
                },
                "limit_rows": {
                    "type": "integer",
                    "description": "可选：限制每个 sheet 返回的最大数据行数",
                },
                "drop_empty": {
                    "type": "boolean",
                    "description": "可选：是否过滤掉空行，默认 true",
                },
            },
            "required": ["path"],
            "callable": read_excel,
        },
        "read_csv": {
            "description": "读取 CSV 文件并返回结构化结果",
            "parameters": {
                "path": {"type": "string", "description": "CSV 文件路径"},
                "encoding": {
                    "type": "string",
                    "description": "可选：文件编码，默认 utf-8",
                },
                "delimiter": {
                    "type": "string",
                    "description": "可选：分隔符，默认逗号",
                },
                "limit_rows": {
                    "type": "integer",
                    "description": "可选：限制返回的最大数据行数",
                },
            },
            "required": ["path"],
            "callable": read_csv,
        },
    }


def _normalize_header(raw_headers: List[Optional[str]]) -> List[str]:
    seen = {}
    normalized = []
    for idx, header in enumerate(raw_headers):
        base_name = str(header).strip() if header not in (None, "") else f"column_{idx + 1}"
        name = base_name
        counter = 2
        while name in seen:
            name = f"{base_name}_{counter}"
            counter += 1
        seen[name] = True
        normalized.append(name)
    return normalized


def _sheet_to_records(ws, limit_rows: Optional[int], drop_empty: bool) -> Dict[str, List[dict]]:
    rows = list(ws.values)
    if not rows:
        return {"columns": [], "rows": [], "row_count": 0}

    headers = _normalize_header(list(rows[0]))
    payload = []

    for values in rows[1:]:
        row_values = list(values or [])
        if drop_empty and all(v in (None, "") for v in row_values):
            continue
        row_dict = {headers[i]: row_values[i] if i < len(row_values) else None for i in range(len(headers))}
        payload.append(row_dict)
        if limit_rows is not None and len(payload) >= limit_rows:
            break

    return {
        "columns": headers,
        "rows": payload,
        "row_count": len(payload),
    }


def read_excel(path: str, sheets: Optional[List[str]] = None, limit_rows: Optional[int] = None,
               drop_empty: bool = True) -> dict:
    if not os.path.exists(path):
        return {"error": f"Excel 文件不存在: {path}"}

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_names = sheets or wb.sheetnames
    result_sheets = []

    for name in sheet_names:
        if name not in wb.sheetnames:
            result_sheets.append({
                "name": name,
                "error": "该工作表不存在",
            })
            continue

        ws = wb[name]
        sheet_payload = _sheet_to_records(ws, limit_rows, drop_empty)
        sheet_payload.update({
            "name": name,
        })
        result_sheets.append(sheet_payload)

    return {
        "path": os.path.abspath(path),
        "sheet_count": len(result_sheets),
        "sheets": result_sheets,
    }


def read_csv(path: str, encoding: str = "utf-8", delimiter: str = ",",
             limit_rows: Optional[int] = None) -> dict:
    if not os.path.exists(path):
        return {"error": f"CSV 文件不存在: {path}"}

    rows = []
    with open(path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        columns = reader.fieldnames or []
        for row in reader:
            rows.append(row)
            if limit_rows is not None and len(rows) >= limit_rows:
                break

    return {
        "path": os.path.abspath(path),
        "columns": columns,
        "row_count": len(rows),
        "rows": rows,
    }
